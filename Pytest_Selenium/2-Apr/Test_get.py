import openpyxl
import pytest
from selenium.webdriver import Chrome,ChromeOptions
from selenium.webdriver.common.by import By

wb = openpyxl.Workbook()
sheetName = "Sheet1"
if sheetName in wb.sheetnames:
    ws = wb[sheetName]
else:
    ws = wb.create_sheet(sheetName)

ws['A1'] ='user'
ws['B1']='password'

#wb.save('sample.xlsx') #it will save xl file in local directory

ws.append(['user1','123'])
ws.append(['akshay','123'])
ws.append(['akshay','12345'])
ws.append(['aks','1'])


wb.save('sample2.xlsx')

def get_data():
    wb = openpyxl.load_workbook("sample2.xlsx")
    ws = wb["Sheet1"]

    data = []

    for row in ws.iter_rows(min_row=2,values_only=True):
        data.append(row)

    return data

@pytest.fixture(autouse=True)
def setup():
    #this is setup
    o = ChromeOptions()
    o.add_experimental_option("detach", True)
    global driver
    driver = Chrome(options=o)
    driver.implicitly_wait(20)
    driver.get('https://www.saucedemo.com/')
    driver.maximize_window()
    print("Open Browser")
    yield
    #teardown
    driver.quit()
    print("\nClose Browser")

@pytest.mark.parametrize('un,passwd',get_data())
def test_login(un,passwd):
    driver.find_element(By.ID,"user-name").send_keys(un)
    driver.find_element(By.ID,"password").send_keys(passwd)
    driver.find_element(By.ID,"login-button").click()
    expected = "https://www.saucedemo.com/inventory.html"
    current_url = driver.current_url
    assert current_url == expected, "Login Failed"