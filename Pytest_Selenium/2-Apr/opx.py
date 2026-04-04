import pytest
import openpyxl

wb = openpyxl.Workbook()
sheetName = "sheet1"
if sheetName in wb.sheetnames:
    ws = wb[sheetName]
else:
    ws = wb.create_sheet(sheetName)

ws['A1'] ='user'
ws['B1']='password'

#wb.save('sample.xlsx') #it will save xl file in local directory

ws.append(['user1','123'])
ws.append(['akshay','123'])
ws.append(['akshay','12345'])
ws.append(['aks','1'])


wb.save('sample2.xlsx')

# def func1():
#     wb = openpyxl.load_workbook("sample.xlsx")
#     ws = wb["Sheet1"]
#
#     data = []
#
#     for row in ws.iter_rows(min_row=2,values_only=True):
#         data.append(row)
#
#     return data